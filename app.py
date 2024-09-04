import streamlit as st
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.styles import numbers
import io
from streamlit_option_menu import option_menu

# Function to normalize percentage values
def normalize_percentage(value):
    if pd.isna(value):
        return value
    value_str = str(value).replace('%', '').strip()
    try:
        return float(value_str)
    except ValueError:
        return value

# Function to replace incorrect values with correct ones and highlight the corrected cells
def replace_and_highlight_cells(ws, merged_df, subset_df1, df1):
    for index, row in merged_df.iterrows():
        supplier_number = row["Supplier Number"]

        # Find the corresponding row in the original df1 DataFrame
        original_index = subset_df1.index[subset_df1["Supplier Number"] == supplier_number].tolist()
        if not original_index:
            continue
        
        original_index = original_index[0]
        cell_row = original_index + 2  # +2 to account for header and 0-based index
        
        for col in ["Score Card", "Expiry Date", "B-BBEE Status", "Black Designated %", "Black Owned %", "Black Women Owned %"]:
            incorrect_col = col + "_incorrect"
            correct_col = col + "_correct"
            cell_col = df1.columns.get_loc(col) + 1

            incorrect_value = row[incorrect_col]
            correct_value = row[correct_col]

            if pd.notna(correct_value):
                if pd.notna(incorrect_value):
                    if isinstance(incorrect_value, float) and isinstance(correct_value, float):
                        if not (abs(incorrect_value - correct_value) < 0.1):
                            ws.cell(row=cell_row, column=cell_col, value=correct_value)
                            ws.cell(row=cell_row, column=cell_col).fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
                    else:
                        if incorrect_value != correct_value:
                            ws.cell(row=cell_row, column=cell_col, value=correct_value)
                            ws.cell(row=cell_row, column=cell_col).fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
                else:
                    ws.cell(row=cell_row, column=cell_col, value=correct_value)
                    ws.cell(row=cell_row, column=cell_col).fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")

# Function to format SAP input file
def format_sap_input(file):
    df = pd.read_excel(file)

    required_columns = [
        'Year', 'Supplier', 'Supplier Description', 'Division', 'Spend Exclusion Percentage', 
        'Supplier Development', 'Scorecard', 'Black Ownership Status', 'Black Woman Ownership Status', 
        'QSE', 'EME', 'Level', 'Black Ownership Percentage', 'Black Woman Ownership Percentage', 
        'Empowering Supplier', 'Black Designated Group', 'Black Designated Group Percentage', 
        'Expiry Date', 'Vat Registration Number', 'File Path'
    ]

    df.rename(columns={
        'Vendor': 'Supplier',
        'Name': 'Supplier Description',
        'POSpend': 'Spend Exclusion Percentage',
        'SuppDev': 'Supplier Development',
        'ScoreCard': 'Scorecard',
        'BOW': 'Black Ownership Status',
        'WOW': 'Black Woman Ownership Status',
        'BOWP': 'Black Ownership Percentage',
        'WOWP': 'Black Woman Ownership Percentage',
        'EMP': 'Empowering Supplier',
        'BDG': 'Black Designated Group',
        'BDGP': 'Black Designated Group Percentage',
        'Vat Reg no': 'Vat Registration Number',
        'Certificate Path': 'File Path'
    }, inplace=True)

    df = df[required_columns]

    # Clean and convert percentage columns
    for col in ['Black Ownership Percentage', 'Black Woman Ownership Percentage', 'Black Designated Group Percentage']:
        # Remove any non-numeric characters, like '%', and convert to numeric, coercing errors
        df[col] = pd.to_numeric(df[col].astype(str).str.replace('[^0-9.]', '', regex=True), errors='coerce')

        # Divide by 100 if the value is greater than 100 to correct the scale
        df[col] = df[col].apply(lambda x: x / 100 if pd.notna(x) and x > 100 else x).astype(str).str[:6]

    df['Year'] = df['Year'].astype(str).str[:4]
    df['Supplier'] = df['Supplier'].astype(str)
    df['Supplier'] = df['Supplier'].apply(lambda x: '0000000' + x if len(x) == 3 and x.isdigit() else ('00' + x if len(x) == 8 and x.isdigit() else x))
    df['Supplier Description'] = df['Supplier Description'].astype(str).str[:35]
    df['Division'] = df['Division'].astype(str).str[:8]
    df['Spend Exclusion Percentage'] = df['Spend Exclusion Percentage'].astype(str).str[:3]
    df['Supplier Development'] = df['Supplier Development'].astype(str).str[:1]
    df['Scorecard'] = df['Scorecard'].replace('Gen', 'GEN').astype(str).str[:3]
    df['Black Ownership Status'] = df['Black Ownership Status'].astype(str).str[:1]
    df['Black Woman Ownership Status'] = df['Black Woman Ownership Status'].astype(str).str[:1]
    df['QSE'] = df['QSE'].astype(str).str[:1]
    df['EME'] = df['EME'].astype(str).str[:1]
    df['Level'] = df['Level'].replace('Non-compliant', '').astype(str).str[:1]
    df['Empowering Supplier'] = df['Empowering Supplier'].astype(str).str[:1]
    df['Black Designated Group'] = df['Black Designated Group'].astype(str).str[:1]
    df['Expiry Date'] = df['Expiry Date'].astype(str).str[:8]
    df['Vat Registration Number'] = df['Vat Registration Number'].astype(str).str[:20]
    df['File Path'] = df['File Path'].astype(str).str[:128]

    df.replace(['n', 'nan', 'NaN'], '', inplace=True)
    df.fillna('', inplace=True)

    # Create an in-memory buffer to save the full formatted file
    buffer_full = io.BytesIO()
    with pd.ExcelWriter(buffer_full, engine='openpyxl') as writer:
        df.to_excel(writer, index=False)
        wb = writer.book
        ws = wb.active

        # Format the Supplier column as text
        for cell in ws['B']:
            cell.number_format = numbers.FORMAT_TEXT

        tab = Table(displayName="FormattedData", ref=ws.dimensions)
        style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
                               showLastColumn=False, showRowStripes=True, showColumnStripes=True)
        tab.tableStyleInfo = style
        ws.add_table(tab)
    buffer_full.seek(0)

    return buffer_full

# Set the custom page title
st.set_page_config(page_title="B-BBEE Data Processing")

st.title('B-BBEE Data Processing')

# Sidebar menu
with st.sidebar:
    menu = option_menu(
        menu_title="BBBEE Excel Data Processing", 
        options=["Validate SAP Data", "Format SAP Input File"], 
        icons=["check-circle", "file-earmark-spreadsheet"], 
        menu_icon="house", 
        default_index=0,
        styles={
            "container": {"padding": "5!important", "background-color": "#262730"},
            "icon": {"color": "white", "font-size": "18px"}, 
            "nav-link": {"font-size": "14px", "text-align": "left", "margin": "0px", "--hover-color": "#565656", "color": "white"},
            "nav-link-selected": {"background-color": "#4CAF50"},
            "menu-title": {"font-size": "16px", "color": "white", "text-align": "left", "white-space": "nowrap"}
        }
    )

if menu == "Validate SAP Data":
    st.header('Validate SAP Data')
    uploaded_file1 = st.file_uploader("Upload Excel From SAP", type=["xlsx"])
    uploaded_file2 = st.file_uploader("Upload Excel With Mpowered Data", type=["xlsx"])

    if uploaded_file1 and uploaded_file2:
        with st.spinner('Processing...'):
            df1 = pd.read_excel(uploaded_file1)
            df2 = pd.read_excel(uploaded_file2)

            # Columns to compare and correct
            cols_df1 = ["Supplier Number", "Score Card", "Expiry Date", "B-BBEE Status", "Black Designated %", "Black Owned %", "Black Women Owned %"]
            cols_df2 = ["Vendor Code*", "Vendor Size", "Expiry Date (dd/mm/yyyy)", "Level", "% Black Designated Group Owned", "% Black Owned", "% Black Women Owned"]

            # Subset the dataframes to only the relevant columns
            subset_df1 = df1[cols_df1].copy()
            subset_df2 = df2[cols_df2].copy()

            # Rename columns in the second dataframe to match the first dataframe for easier merging
            subset_df2.columns = ["Supplier Number", "Score Card", "Expiry Date", "B-BBEE Status", "Black Designated %", "Black Owned %", "Black Women Owned %"]

            # Normalize percentage columns
            percentage_cols = ["Black Designated %", "Black Owned %", "Black Women Owned %"]
            for col in percentage_cols:
                subset_df1[col] = subset_df1[col].apply(normalize_percentage)
                subset_df2[col] = subset_df2[col].apply(normalize_percentage)

            # Merge the dataframes on Supplier Number
            merged_df = pd.merge(subset_df1, subset_df2, on="Supplier Number", suffixes=('_incorrect', '_correct'))

            # Load the original Excel file to apply the highlighting and corrections
            wb = load_workbook(uploaded_file1)
            ws = wb.active

            replace_and_highlight_cells(ws, merged_df, subset_df1, df1)
            
            # Save the updated Excel file to a BytesIO object
            corrected_file = io.BytesIO()
            wb.save(corrected_file)
            corrected_file.seek(0)

        st.success('The values have been successfully corrected.')
        st.download_button('Download Corrected File', data=corrected_file, file_name='Corrected_Spend_Report.xlsx')

elif menu == "Format SAP Input File":
    st.header('Format SAP Input File')
    uploaded_file = st.file_uploader("Upload SAP Input Excel file", type=["xlsx"])

    if uploaded_file:
        with st.spinner('Processing...'):
            buffer_full = format_sap_input(uploaded_file)
        
        st.success('The SAP input file has been successfully formatted.')
        st.download_button('Download Formatted File', data=buffer_full, file_name='Formatted_Input_to_SAP_June_2024.xlsx')
